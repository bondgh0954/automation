import openpyxl as op

work_book = op.load_workbook('inventory.xlsx')
work_sheet = work_book['Sheet1']

supplier_row_dictionary = {}
inventory_price = {}
inventory_less = {}

for supplier in range(2, work_sheet.max_row + 1):
    supplier_list = work_sheet.cell(supplier, 4).value
    price = work_sheet.cell(supplier, 3).value
    inventory = work_sheet.cell(supplier, 2).value
    product_number = work_sheet.cell(supplier, 1).value

    if supplier_list in supplier_row_dictionary:
        supplier_name = supplier_row_dictionary[supplier_list]
        supplier_row_dictionary[supplier_list] = supplier_name + 1
    else:
        supplier_row_dictionary[supplier_list] = 1

    if supplier_list in inventory_price:
        total_inventory = inventory_price[supplier_list]
        inventory_price[supplier_list] = total_inventory + price * inventory
    else:
        inventory_price[supplier_list] = price * inventory

    if inventory < 10:
        inventory_less[product_number] = inventory

print(inventory_less)
print(supplier_row_dictionary)
print(inventory_price)
